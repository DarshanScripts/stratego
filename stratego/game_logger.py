# stratego/game_logger.py
# [ENHANCED - 21 Jan 2026] Major upgrade to support comprehensive game metrics
# Added: move timing, winner/loser tracking, invalid move counting, repetition detection,
# game end reason tracking, and generation of Excel reports with multiple sheets
# [RESTRUCTURED - 21 Jan 2026] Changed to generate:
#   1. One CSV per game with all details
#   2. One Master Excel with 2 sheets (All Games Results + Summary Statistics)
from __future__ import annotations
import csv
import time
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Install with: pip install openpyxl")

class GameLogger:
    """
    Enhanced GameLogger that tracks game metrics and move times.
    Creates CSV files per game with detailed move-by-move logs and summary metrics.
    """
    def __init__(self, out_dir: str, game_id: Optional[str] = None, prompt_name: str = "", game_type: str = "standard", board_size: int = 10):
        # Setup directories
        self.out_dir = out_dir
        self.logs_dir = Path(out_dir)
        self.games_dir = self.logs_dir / "games"
        self.games_dir.mkdir(parents=True, exist_ok=True)
        
        # Game identification
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.game_id = game_id or ts
        self.prompt_name = prompt_name
        self.game_type = game_type
        self.board_size = board_size
        
        # [ENHANCED - 21 Jan 2026] Added comprehensive metrics tracking for game analysis
        # Tracks: winner/loser, game end reason, flag capture, turn counts, invalid moves,
        # repetitions, move history, timing data, and game duration
        self.player_0_name = ""
        self.player_1_name = ""
        self.winner: Optional[int] = None
        self.loser: Optional[int] = None
        self.game_end_reason = ""
        self.flag_captured_by: Optional[int] = None
        self.total_turns = 0
        self.player_0_turns = 0
        self.player_1_turns = 0
        self.player_0_invalid_moves = 0
        self.player_1_invalid_moves = 0
        self.repetition_moves = 0
        self.move_history: List[str] = []
        self.move_times: List[dict] = []
        self.start_time = time.time()
        self.end_time: Optional[float] = None
        
        # CSV file for move-by-move logging
        self.path = os.path.join(self.games_dir, f"{self.game_id}.csv")
        self._f = open(self.path, "w", newline="", encoding="utf-8")
        self._writer = csv.DictWriter(
            self._f,
            fieldnames=[
                "turn", "player", "model_name",
                "move", "from_pos", "to_pos", "piece_type",
                "move_direction", "target_piece", "battle_outcome",
                "board_size", "time_taken_seconds", "game_end_reason",
            ],
            quoting=csv.QUOTE_MINIMAL,
            escapechar="\\"
        )
        self._writer.writeheader()
        self._rows: List[dict] = []
    
    def set_players(self, player_0_name: str, player_1_name: str):
        """Set player names for metrics"""
        self.player_0_name = player_0_name
        self.player_1_name = player_1_name
    
    def log_move_start(self, player: int, turn: int):
        """[NEW - 21 Jan 2026] Track move start time for performance analysis
        Returns: timestamp to be passed to log_move for timing calculation
        """
        return time.time()

    def log_move(
        self,
        turn: int,
        player: int,
        move: str,
        model_name: str = "",
        src: str = "",
        dst: str = "",
        piece_type: str = "",
        outcome: str = "",
        board_state: str = "",
        available_moves: str = "",
        move_direction: str = "",
        target_piece: str = "",
        battle_outcome: str = "",
        start_time: Optional[float] = None,
        is_valid: bool = True,
    ):
        """[ENHANCED - 21 Jan 2026] Extended to track move timing and validity
        Now calculates time taken, tracks repetitions, counts turns/invalid moves,
        and stores timing data for each player's performance analysis
        """
        # Calculate time taken
        time_taken = time.time() - start_time if start_time else 0.0
        
        # Track metrics
        self.move_history.append(move)
        if self.move_history.count(move) >= 3:
            self.repetition_moves += 1
        
        # Store move time data
        self.move_times.append({
            'player': player,
            'turn': turn,
            'time_taken': time_taken,
            'is_valid': is_valid,
            'model_name': model_name
        })
        
        # Update turn counters
        if is_valid:
            if player == 0:
                self.player_0_turns += 1
            else:
                self.player_1_turns += 1
            self.total_turns += 1
        else:
            if player == 0:
                self.player_0_invalid_moves += 1
            else:
                self.player_1_invalid_moves += 1
        
        # Store player names if not set
        if model_name:
            if player == 0 and not self.player_0_name:
                self.player_0_name = model_name
            elif player == 1 and not self.player_1_name:
                self.player_1_name = model_name
        
        # Prepare row data (streamlined columns only)
        row = {
            "turn": turn,
            "player": player,
            "model_name": model_name,
            "move": move,
            "from_pos": src,
            "to_pos": dst,
            "piece_type": piece_type,
            "move_direction": move_direction,
            "target_piece": target_piece,
            "battle_outcome": battle_outcome,
            "board_size": self.board_size,
            "time_taken_seconds": round(time_taken, 4),
            "game_end_reason": "",  # Will be filled in last row only
        }
        
        # Store and write
        self._rows.append(row)
        try:
            self._writer.writerow(row)
            self._f.flush()
        except Exception as e:
            print(f"❌ ERROR writing move to CSV: {e}")
            print(f"   Row data: {row}")
    
    def log_invalid_move(self, player: int):
        """[NEW - 21 Jan 2026] Track invalid moves per player for analysis"""
        if player == 0:
            self.player_0_invalid_moves += 1
        else:
            self.player_1_invalid_moves += 1
    
    def set_game_end(self, winner: int, reason: str, flag_captured: bool = False):
        """[NEW - 21 Jan 2026] Record game end details including winner, reason, and flag capture status"""
        self.end_time = time.time()
        self.winner = winner
        self.loser = 1 - winner if winner in [0, 1] else None
        self.game_end_reason = reason
        
        if flag_captured:
            self.flag_captured_by = winner
    
    def finalize_game(self, winner: Optional[int] = None, game_result: str = ""):
        """[ENHANCED - 21 Jan 2026] Finalize game and generate reports
        Creates:
        1. One CSV with all game details
        2. Updates Master Excel with 2 sheets (All Games + Summary Statistics)
        """
        if winner is None:
            winner = self.winner
        
        if game_result and not self.game_end_reason:
            self.game_end_reason = game_result
        
        # Close move log file
        self._f.close()
        
        # Update ONLY the last row with game end information
        if self._rows:
            winner_name = self.player_0_name if winner == 0 else (self.player_1_name if winner == 1 else "Draw")
            # Include flag capture status in the reason text
            flag_suffix = " - Flag Captured" if hasattr(self, 'flag_captured_by') and self.flag_captured_by is not None else ""
            game_end_info = f"{winner_name} wins - {self.game_end_reason or game_result}{flag_suffix}"
            self._rows[-1]["game_end_reason"] = game_end_info
        
        # [NEW - 21 Jan 2026] Save detailed per-game CSV with all move data
        self._save_detailed_game_csv()
        
        # [NEW - 21 Jan 2026] Update Master Excel with both sheets
        print(f"🔧 DEBUG: EXCEL_AVAILABLE = {EXCEL_AVAILABLE}")
        if EXCEL_AVAILABLE:
            try:
                print(f"🔧 DEBUG: Calling _update_master_excel()...")
                self._update_master_excel()
                print(f"✅ DEBUG: _update_master_excel() completed successfully")
            except Exception as e:
                print(f"❌ ERROR updating Master Excel: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("⚠️  Excel reports skipped - openpyxl not installed")
    
    def _save_detailed_game_csv(self):
        """[NEW - 21 Jan 2026] Save comprehensive CSV for this game only"""
        with open(self.path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "turn", "player", "model_name",
                    "move", "from_pos", "to_pos", "piece_type",
                    "move_direction", "target_piece", "battle_outcome",
                    "board_size", "time_taken_seconds", "game_end_reason",
                ],
                quoting=csv.QUOTE_MINIMAL,
                escapechar="\\"
            )
            writer.writeheader()
            writer.writerows(self._rows)
        
        print(f"📊 Game details saved to {self.path}")
    
    def _update_master_excel(self):
        """[NEW - 21 Jan 2026] Rebuild Master Excel from ALL CSV files in logs/games
        Sheet 1: All Games Results (one row per game)
        Sheet 2: Summary Statistics (aggregated model comparison)
        """
        master_file = self.games_dir / "Master_Game_Results.xlsx"
        
        # [CRITICAL - 21 Jan 2026] Read ALL CSV files from logs/games to rebuild master
        all_games_data = []
        csv_files = sorted([f for f in self.games_dir.glob("*.csv")])
        
        for csv_file in csv_files:
            try:
                with open(csv_file, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                    if not rows:
                        continue
                    
                    # Extract game-level data from first row (all rows have same game data)
                    first_row = rows[0]
                    
                    # Extract model names for each player
                    model_p0 = 'Unknown'
                    model_p1 = 'Unknown'
                    for r in rows:
                        if r.get('player') == '0' and r.get('model_name'):
                            model_p0 = r.get('model_name')
                            break
                    for r in rows:
                        if r.get('player') == '1' and r.get('model_name'):
                            model_p1 = r.get('model_name')
                            break
                    
                    # Count metrics across all rows
                    total_turns = len([r for r in rows if r.get('player') in ['0', '1']])
                    
                    # Note: We no longer track invalid moves in CSV (removed is_valid column)
                    # These will be 0 for all games with new format
                    invalid_p0 = 0
                    invalid_p1 = 0
                    
                    # Count repetitions per player (moves with same from_pos -> to_pos)
                    moves_seen_p0 = set()
                    moves_seen_p1 = set()
                    repetitions_p0 = 0
                    repetitions_p1 = 0
                    for r in rows:
                        move_key = f"{r.get('from_pos')}->{r.get('to_pos')}"
                        if move_key != "None->None" and move_key != "->" and r.get('player'):
                            if r.get('player') == '0':
                                if move_key in moves_seen_p0:
                                    repetitions_p0 += 1
                                moves_seen_p0.add(move_key)
                            elif r.get('player') == '1':
                                if move_key in moves_seen_p1:
                                    repetitions_p1 += 1
                                moves_seen_p1.add(move_key)
                    
                    # Extract winner from game_end_reason (last row)
                    # Format: "qwen3:8b wins - Player 1 wins" or "Draw - Stalemate"
                    last_row = rows[-1] if rows else first_row
                    game_end_reason = last_row.get('game_end_reason', '').strip()
                    
                    # If game_end_reason is empty, set to Unknown
                    if not game_end_reason:
                        game_end_reason = 'Unknown'
                    
                    winner = -1  # Default to draw
                    if game_end_reason and game_end_reason != 'Unknown':
                        reason_lower = game_end_reason.lower()
                        if 'player 1 wins' in reason_lower or model_p0 in game_end_reason:
                            winner = 0
                        elif 'player 2 wins' in reason_lower or model_p1 in game_end_reason:
                            winner = 1
                        elif 'draw' in reason_lower or 'stalemate' in reason_lower:
                            winner = -1
                    
                    # Calculate time taken per player
                    time_p0 = []
                    time_p1 = []
                    for r in rows:
                        try:
                            t = float(r.get('time_taken_seconds', 0))
                            if t > 0:
                                if r.get('player') == '0':
                                    time_p0.append(t)
                                elif r.get('player') == '1':
                                    time_p1.append(t)
                        except:
                            pass
                    
                    total_time_p0 = round(sum(time_p0), 2) if time_p0 else 0
                    total_time_p1 = round(sum(time_p1), 2) if time_p1 else 0
                    
                    # game_end_reason already extracted above at line 312
                    # Don't re-extract, just use the existing value
                    
                    # [NEW - 22 Jan 2026] Parse game end reason to extract specific end conditions
                    reason_lower = game_end_reason.lower()
                    flag_captured = 1 if 'flag' in reason_lower and 'captured' in reason_lower else 0
                    max_turns_reached = 1 if 'max turns' in reason_lower or 'turn limit' in reason_lower else 0
                    no_valid_moves = 1 if 'no valid moves' in reason_lower or 'no moves' in reason_lower else 0
                    stalemate = 1 if 'stalemate' in reason_lower and 'no valid moves' not in reason_lower else 0
                    
                    game_data = {
                        'model_p0': model_p0,
                        'model_p1': model_p1,
                        'board_size': int(first_row.get('board_size', 10)),
                        'winner': winner,
                        'turns': total_turns,
                        'invalid_moves_p0': invalid_p0,
                        'invalid_moves_p1': invalid_p1,
                        'num_invalid_moves': invalid_p0 + invalid_p1,
                        'repetitions_p0': repetitions_p0,
                        'repetitions_p1': repetitions_p1,
                        'time_p0': total_time_p0,
                        'time_p1': total_time_p1,
                        'flag_captured': flag_captured,
                        'max_turns_reached': max_turns_reached,
                        'no_valid_moves': no_valid_moves,
                        'stalemate': stalemate,
                        'game_end_reason': game_end_reason,
                    }
                    all_games_data.append(game_data)
                    
            except Exception as e:
                print(f"⚠️  Warning: Could not read {csv_file.name}: {e}")
                continue
        
        # Create new workbook (rebuild from scratch each time)
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # [SHEET 1 - 21 Jan 2026] All Games Results
        ws1 = wb.create_sheet("All Games Results")
        
        # Add headers with styling
        if all_games_data:
            headers = list(all_games_data[0].keys())
            ws1.append(headers)
            for cell in ws1[1]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add all game data
            for game_data in all_games_data:
                ws1.append(list(game_data.values()))
        
        # [SHEET 2 - 22 Jan 2026] Matchup Comparison (Player 0 vs Player 1)
        self._create_matchup_comparison_sheet(wb, all_games_data)
        
        # [SHEET 3 - 22 Jan 2026] Summary Statistics (Per-Matchup)
        self._update_summary_statistics_sheet_matchup_based(wb, all_games_data)
        
        # [SHEET 4 - 22 Jan 2026] Charts (Per-Matchup)
        self._create_charts_sheet_matchup_based(wb, all_games_data)
        
        # Save workbook
        wb.save(master_file)
        print(f"📊 Master Excel updated: {master_file}")
    
    def _create_matchup_comparison_sheet(self, wb, all_games_data):
        """[NEW - 22 Jan 2026] Create matchup-based comparison (P0 vs P1)
        Shows head-to-head results for each model pairing
        """
        if "Matchup Comparison" in wb.sheetnames:
            wb.remove(wb["Matchup Comparison"])
        
        ws = wb.create_sheet("Matchup Comparison", 0)  # Insert as first sheet
        
        if not all_games_data:
            return
        
        # Group games by matchup (model_p0 vs model_p1)
        matchups = {}
        for game in all_games_data:
            model_p0 = str(game.get('model_p0', 'Unknown'))
            model_p1 = str(game.get('model_p1', 'Unknown'))
            matchup_key = f"{model_p0} vs {model_p1}"
            
            if matchup_key not in matchups:
                matchups[matchup_key] = {
                    'model_p0': model_p0,
                    'model_p1': model_p1,
                    'total_games': 0,
                    'p0_wins': 0,
                    'p1_wins': 0,
                    'draws': 0,
                    'total_turns': [],
                    'p0_time': [],
                    'p1_time': [],
                    'p0_repetitions': [],
                    'p1_repetitions': [],
                    'flag_captured': 0,
                    'no_moves': 0,
                    'stalemate': 0,
                    'max_turns': 0
                }
            
            m = matchups[matchup_key]
            m['total_games'] += 1
            m['total_turns'].append(game.get('turns', 0))
            m['p0_time'].append(game.get('time_p0', 0))
            m['p1_time'].append(game.get('time_p1', 0))
            m['p0_repetitions'].append(game.get('repetitions_p0', 0))
            m['p1_repetitions'].append(game.get('repetitions_p1', 0))
            
            winner = game.get('winner', -1)
            if winner == 0:
                m['p0_wins'] += 1
            elif winner == 1:
                m['p1_wins'] += 1
            else:
                m['draws'] += 1
            
            # [UPDATED - 22 Jan 2026] Use dedicated columns for end reason tracking
            m['flag_captured'] += game.get('flag_captured', 0)
            m['no_moves'] += game.get('no_valid_moves', 0)
            m['stalemate'] += game.get('stalemate', 0)
            m['max_turns'] += game.get('max_turns_reached', 0)
        
        # Create headers
        headers = [
            'Matchup',
            'Model (P0)',
            'Model (P1)',
            'Total Games',
            'P0 Wins',
            'P1 Wins',
            'Draws',
            'P0 Win %',
            'P1 Win %',
            'Avg Turns',
            'Avg Time P0 (s)',
            'Avg Time P1 (s)',
            'Avg Repetitions P0',
            'Avg Repetitions P1',
            'Flag Captured',
            'No Moves',
            'Stalemate',
            'Max Turns Reached'
        ]
        ws.append(headers)
        
        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add matchup data
        for matchup_key in sorted(matchups.keys()):
            m = matchups[matchup_key]
            
            avg_turns = round(sum(m['total_turns']) / len(m['total_turns']), 1) if m['total_turns'] else 0
            avg_time_p0 = round(sum(m['p0_time']) / len(m['p0_time']), 2) if m['p0_time'] else 0
            avg_time_p1 = round(sum(m['p1_time']) / len(m['p1_time']), 2) if m['p1_time'] else 0
            avg_reps_p0 = round(sum(m['p0_repetitions']) / len(m['p0_repetitions']), 1) if m['p0_repetitions'] else 0
            avg_reps_p1 = round(sum(m['p1_repetitions']) / len(m['p1_repetitions']), 1) if m['p1_repetitions'] else 0
            
            p0_win_pct = round((m['p0_wins'] / m['total_games']) * 100, 1) if m['total_games'] > 0 else 0
            p1_win_pct = round((m['p1_wins'] / m['total_games']) * 100, 1) if m['total_games'] > 0 else 0
            
            row = [
                matchup_key,
                m['model_p0'],
                m['model_p1'],
                m['total_games'],
                m['p0_wins'],
                m['p1_wins'],
                m['draws'],
                f"{p0_win_pct}%",
                f"{p1_win_pct}%",
                avg_turns,
                avg_time_p0,
                avg_time_p1,
                avg_reps_p0,
                avg_reps_p1,
                m['flag_captured'],
                m['no_moves'],
                m['stalemate'],
                m['max_turns']
            ]
            ws.append(row)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
            ws.column_dimensions[col].width = 12
    
    def _update_summary_statistics_sheet_matchup_based(self, wb, all_games_data):
        """[NEW - 22 Jan 2026] Generate Summary Statistics per MATCHUP"""
        if "Summary Statistics" in wb.sheetnames:
            wb.remove(wb["Summary Statistics"])
        
        ws2 = wb.create_sheet("Summary Statistics")
        
        if not all_games_data:
            return
        
        # Group by matchup
        matchups = {}
        for game in all_games_data:
            model_p0 = str(game.get('model_p0', 'Unknown'))
            model_p1 = str(game.get('model_p1', 'Unknown'))
            matchup_key = f"{model_p0} vs {model_p1}"
            
            if matchup_key not in matchups:
                matchups[matchup_key] = {
                    'games': 0,
                    'p0_wins': 0,
                    'p1_wins': 0,
                    'draws': 0,
                    'total_turns': [],
                    'flag_captured': 0,
                    'max_turns': 0,
                    'no_moves': 0,
                    'stalemate': 0
                }
            
            m = matchups[matchup_key]
            m['games'] += 1
            m['total_turns'].append(game.get('turns', 0))
            m['flag_captured'] += game.get('flag_captured', 0)
            m['max_turns'] += game.get('max_turns_reached', 0)
            m['no_moves'] += game.get('no_valid_moves', 0)
            m['stalemate'] += game.get('stalemate', 0)
            
            winner = game.get('winner', -1)
            if winner == 0:
                m['p0_wins'] += 1
            elif winner == 1:
                m['p1_wins'] += 1
            else:
                m['draws'] += 1
        
        # Header
        ws2.cell(1, 1, "📊 MATCHUP STATISTICS").font = Font(bold=True, size=14)
        ws2.cell(1, 1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Table headers
        headers = ["Matchup", "Games", "P0 Wins", "P1 Wins", "Draws", "Avg Turns", "Flag Captured", "Max Turns", "No Moves", "Stalemate"]
        ws2.append([])
        ws2.append(headers)
        for cell in ws2[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for matchup_key in sorted(matchups.keys()):
            m = matchups[matchup_key]
            avg_turns = round(sum(m['total_turns']) / len(m['total_turns']), 1) if m['total_turns'] else 0
            
            ws2.append([
                matchup_key,
                m['games'],
                m['p0_wins'],
                m['p1_wins'],
                m['draws'],
                avg_turns,
                m['flag_captured'],
                m['max_turns'],
                m['no_moves'],
                m['stalemate']
            ])
        
        # Column widths
        ws2.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws2.column_dimensions[col].width = 15
    
    def _create_charts_sheet_matchup_based(self, wb, all_games_data):
        """[NEW - 22 Jan 2026] Create Charts per MATCHUP"""
        try:
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList
        except ImportError:
            print("⚠️  Chart creation requires openpyxl with chart support")
            return
        
        if "Charts" in wb.sheetnames:
            wb.remove(wb["Charts"])
        
        ws3 = wb.create_sheet("Charts")
        
        if not all_games_data:
            return
        
        # Group by matchup
        matchups = {}
        for game in all_games_data:
            model_p0 = str(game.get('model_p0', 'Unknown'))
            model_p1 = str(game.get('model_p1', 'Unknown'))
            matchup_key = f"{model_p0} vs {model_p1}"
            
            if matchup_key not in matchups:
                matchups[matchup_key] = {
                    'p0_wins': 0,
                    'p1_wins': 0,
                    'draws': 0,
                    'avg_turns': [],
                    'flag_captured': 0
                }
            
            m = matchups[matchup_key]
            m['avg_turns'].append(game.get('turns', 0))
            m['flag_captured'] += game.get('flag_captured', 0)
            
            winner = game.get('winner', -1)
            if winner == 0:
                m['p0_wins'] += 1
            elif winner == 1:
                m['p1_wins'] += 1
            else:
                m['draws'] += 1
        
        # Data table
        ws3.append(["Matchup", "P0 Wins", "P1 Wins", "Draws", "Avg Turns", "Flag Captured"])
        ws3[1][0].font = Font(bold=True)
        
        for matchup_key in sorted(matchups.keys()):
            m = matchups[matchup_key]
            avg_turns = round(sum(m['avg_turns']) / len(m['avg_turns']), 1) if m['avg_turns'] else 0
            ws3.append([matchup_key, m['p0_wins'], m['p1_wins'], m['draws'], avg_turns, m['flag_captured']])
        
        # Create charts for each matchup
        chart_row = 2
        for matchup_key in sorted(matchups.keys()):
            # Win comparison bar chart
            chart1 = BarChart()
            chart1.type = "col"
            chart1.title = f"{matchup_key} - Win Comparison"
            chart1.y_axis.title = "Number of Wins"
            chart1.x_axis.title = "Result"
            
            data = Reference(ws3, min_col=2, max_col=4, min_row=chart_row, max_row=chart_row)
            chart1.add_data(data, titles_from_data=False)
            
            chart1.dataLabels = DataLabelList()
            chart1.dataLabels.showVal = True
            
            # Position charts in grid
            chart_col = "H" if (chart_row - 2) % 2 == 0 else "P"
            chart_row_pos = ((chart_row - 2) // 2) * 16 + 2
            ws3.add_chart(chart1, f"{chart_col}{chart_row_pos}")
            
            chart_row += 1
        
        # Column widths
        ws3.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws3.column_dimensions[col].width = 15
    
    def _update_summary_statistics_sheet(self, wb, all_games_ws):
        """[NEW - 21 Jan 2026] Generate comprehensive Summary Statistics sheet"""
        if "Summary Statistics" in wb.sheetnames:
            wb.remove(wb["Summary Statistics"])
        
        ws2 = wb.create_sheet("Summary Statistics")
        
        # Read all game data from Sheet 1
        all_games = []
        headers = [cell.value for cell in all_games_ws[1]]
        for row in all_games_ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            all_games.append(dict(zip(headers, row)))
        
        if not all_games:
            return
        
        total_games = len(all_games)
        
        # Extract unique models (filter None and convert to string)
        all_models = set()
        for game in all_games:
            if game.get('model_p0') is not None:
                all_models.add(str(game['model_p0']))
            if game.get('model_p1') is not None:
                all_models.add(str(game['model_p1']))
        all_models = sorted(list(all_models))
        
        # Calculate comprehensive statistics per model
        model_stats = {model: {
            'total_wins': 0,
            'total_losses': 0,
            'games_played': 0,
            'turns_when_won': [],
            'turns_when_lost': [],
            'invalid_moves_total': 0,
            'repetitions_total': 0,
            'time_taken_total': 0
        } for model in all_models}
        
        draws = 0
        all_turns = []
        all_invalid = []
        all_repetitions = []
        
        for game in all_games:
            winner = game.get('winner')
            model_p0 = str(game.get('model_p0', ''))
            model_p1 = str(game.get('model_p1', ''))
            turns = game.get('turns', 0) if isinstance(game.get('turns'), (int, float)) else 0
            invalid_p0 = game.get('invalid_moves_p0', 0) if isinstance(game.get('invalid_moves_p0'), (int, float)) else 0
            invalid_p1 = game.get('invalid_moves_p1', 0) if isinstance(game.get('invalid_moves_p1'), (int, float)) else 0
            reps_p0 = game.get('repetitions_p0', 0) if isinstance(game.get('repetitions_p0'), (int, float)) else 0
            reps_p1 = game.get('repetitions_p1', 0) if isinstance(game.get('repetitions_p1'), (int, float)) else 0
            num_invalid = game.get('num_invalid_moves', 0) if isinstance(game.get('num_invalid_moves'), (int, float)) else 0
            time_p0 = game.get('time_p0', 0) if isinstance(game.get('time_p0'), (int, float)) else 0
            time_p1 = game.get('time_p1', 0) if isinstance(game.get('time_p1'), (int, float)) else 0
            
            # Skip if model names are invalid
            if not model_p0 or not model_p1 or model_p0 not in model_stats or model_p1 not in model_stats:
                continue
            
            all_turns.append(turns)
            all_invalid.append(num_invalid)
            all_repetitions.append(reps_p0 + reps_p1)
            
            # Update model stats (each model gets its own time and repetitions)
            model_stats[model_p0]['games_played'] += 1
            model_stats[model_p1]['games_played'] += 1
            model_stats[model_p0]['invalid_moves_total'] += invalid_p0
            model_stats[model_p1]['invalid_moves_total'] += invalid_p1
            model_stats[model_p0]['repetitions_total'] += reps_p0
            model_stats[model_p1]['repetitions_total'] += reps_p1
            model_stats[model_p0]['time_taken_total'] += time_p0
            model_stats[model_p1]['time_taken_total'] += time_p1
            
            if winner == 0:
                model_stats[model_p0]['total_wins'] += 1
                model_stats[model_p1]['total_losses'] += 1
                model_stats[model_p0]['turns_when_won'].append(turns)
                model_stats[model_p1]['turns_when_lost'].append(turns)
            elif winner == 1:
                model_stats[model_p1]['total_wins'] += 1
                model_stats[model_p0]['total_losses'] += 1
                model_stats[model_p1]['turns_when_won'].append(turns)
                model_stats[model_p0]['turns_when_lost'].append(turns)
            else:
                draws += 1
        
        # Build comprehensive summary
        row = 1
        
        # Title
        ws2.cell(row, 1, "🎮 STRATEGO AI - COMPREHENSIVE ANALYSIS").font = Font(bold=True, size=14, color="FFFFFF")
        ws2.cell(row, 1).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws2.merge_cells(f"A{row}:D{row}")
        row += 2
        
        # ===== OVERALL STATISTICS =====
        ws2.cell(row, 1, "📊 OVERALL STATISTICS").font = Font(bold=True, size=12)
        ws2.cell(row, 1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        row += 1
        ws2.append(["Total Games Played", total_games])
        ws2.append(["Total Wins (All Models)", sum(m['total_wins'] for m in model_stats.values())])
        ws2.append(["Total Losses (All Models)", sum(m['total_losses'] for m in model_stats.values())])
        ws2.append(["Total Draws", draws])
        ws2.append(["Average Turns per Game", round(sum(all_turns)/len(all_turns), 2) if all_turns else 0])
        ws2.append(["Average Invalid Moves per Game", round(sum(all_invalid)/len(all_invalid), 2) if all_invalid else 0])
        ws2.append(["Average Repetitions per Game", round(sum(all_repetitions)/len(all_repetitions), 2) if all_repetitions else 0])
        row = ws2.max_row + 2
        
        # ===== MODEL PERFORMANCE COMPARISON TABLE =====
        ws2.cell(row, 1, "🏆 MODEL PERFORMANCE COMPARISON").font = Font(bold=True, size=12)
        ws2.cell(row, 1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        row += 1
        
        # Headers
        header_row = ["Model", "Games", "Wins", "Losses", "Win Rate %", "Avg Turns (Win)", "Avg Turns (Loss)", "Avg Invalid/Game", "Avg Time (s)"]
        ws2.append(header_row)
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Model data rows
        for model in all_models:
            stats = model_stats[model]
            games = stats['games_played']
            wins = stats['total_wins']
            losses = stats['total_losses']
            win_rate = round((wins / games * 100), 1) if games > 0 else 0
            avg_turns_win = round(sum(stats['turns_when_won']) / len(stats['turns_when_won']), 1) if stats['turns_when_won'] else 0
            avg_turns_loss = round(sum(stats['turns_when_lost']) / len(stats['turns_when_lost']), 1) if stats['turns_when_lost'] else 0
            avg_invalid = round(stats['invalid_moves_total'] / games, 2) if games > 0 else 0
            avg_time = round(stats['time_taken_total'] / games, 2) if games > 0 else 0
            
            ws2.append([model, games, wins, losses, win_rate, avg_turns_win, avg_turns_loss, avg_invalid, avg_time])
        
        row = ws2.max_row + 2
        
        # ===== DETAILED MODEL BREAKDOWN =====
        ws2.cell(row, 1, "📈 DETAILED MODEL BREAKDOWN").font = Font(bold=True, size=12)
        ws2.cell(row, 1).fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
        row += 1
        
        for model in all_models:
            stats = model_stats[model]
            ws2.append([f"--- {model} ---"])
            ws2[ws2.max_row][0].font = Font(bold=True, size=11)
            ws2.append(["  Total Wins", stats['total_wins']])
            ws2.append(["  Total Losses", stats['total_losses']])
            ws2.append(["  Games Played", stats['games_played']])
            ws2.append(["  Total Invalid Moves", stats['invalid_moves_total']])
            ws2.append(["  Total Repetitions", stats['repetitions_total']])
            ws2.append(["  Total Time Taken (s)", round(stats['time_taken_total'], 2)])
            ws2.append(["  Avg Time per Game (s)", round(stats['time_taken_total'] / stats['games_played'], 2) if stats['games_played'] > 0 else 0])
            ws2.append([])  # Blank row
        
        # Column widths
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 15
        ws2.column_dimensions['F'].width = 18
        ws2.column_dimensions['G'].width = 18
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 15
    
    def _create_charts_sheet(self, wb, all_games_ws):
        """[NEW - 21 Jan 2026] Create Charts & Visualizations sheet with bar and pie charts"""
        try:
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList
        except ImportError:
            print("⚠️  Chart creation requires openpyxl with chart support")
            return
        
        if "Charts" in wb.sheetnames:
            wb.remove(wb["Charts"])
        
        ws3 = wb.create_sheet("Charts")
        
        # Read all game data from Sheet 1
        all_games = []
        headers = [cell.value for cell in all_games_ws[1]]
        for row in all_games_ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            all_games.append(dict(zip(headers, row)))
        
        if not all_games:
            return
        
        # Extract unique models and calculate statistics
        all_models = set()
        for game in all_games:
            if game.get('model_p0'):
                all_models.add(str(game['model_p0']))
            if game.get('model_p1'):
                all_models.add(str(game['model_p1']))
        all_models = sorted(list(all_models))
        
        # Calculate model statistics
        model_wins = {model: 0 for model in all_models}
        model_games = {model: 0 for model in all_models}
        model_avg_turns = {model: [] for model in all_models}
        model_avg_time = {model: [] for model in all_models}
        
        for game in all_games:
            model_p0 = str(game.get('model_p0', ''))
            model_p1 = str(game.get('model_p1', ''))
            winner = game.get('winner')
            turns = game.get('turns', 0) if isinstance(game.get('turns'), (int, float)) else 0
            time_p0 = game.get('time_p0', 0) if isinstance(game.get('time_p0'), (int, float)) else 0
            time_p1 = game.get('time_p1', 0) if isinstance(game.get('time_p1'), (int, float)) else 0
            
            if model_p0 in model_wins:
                model_games[model_p0] += 1
                model_avg_turns[model_p0].append(turns)
                model_avg_time[model_p0].append(time_p0)
                if winner == 0:
                    model_wins[model_p0] += 1
            
            if model_p1 in model_wins:
                model_games[model_p1] += 1
                model_avg_turns[model_p1].append(turns)
                model_avg_time[model_p1].append(time_p1)
                if winner == 1:
                    model_wins[model_p1] += 1
        
        # Calculate averages
        for model in all_models:
            if model_avg_turns[model]:
                model_avg_turns[model] = round(sum(model_avg_turns[model]) / len(model_avg_turns[model]), 1)
            else:
                model_avg_turns[model] = 0
            
            if model_avg_time[model]:
                model_avg_time[model] = round(sum(model_avg_time[model]) / len(model_avg_time[model]), 2)
            else:
                model_avg_time[model] = 0
        
        # === DATA TABLE FOR CHARTS ===
        ws3.append(["Model", "Wins", "Games", "Win Rate %", "Avg Turns", "Avg Time (s)"])
        ws3[1][0].font = Font(bold=True)
        
        for model in all_models:
            games = model_games[model]
            wins = model_wins[model]
            win_rate = round((wins / games * 100), 1) if games > 0 else 0
            ws3.append([model, wins, games, win_rate, model_avg_turns[model], model_avg_time[model]])
        
        # === PIE CHART: Win Distribution ===
        pie_chart = PieChart()
        pie_chart.title = "Win Distribution by Model"
        pie_chart.style = 10
        
        labels = Reference(ws3, min_col=1, min_row=2, max_row=len(all_models) + 1)
        data = Reference(ws3, min_col=2, min_row=1, max_row=len(all_models) + 1)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        
        # Add data labels showing values
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showVal = True
        pie_chart.dataLabels.showCatName = True
        
        ws3.add_chart(pie_chart, "H2")
        
        # === BAR CHART: Average Turns ===
        bar_chart1 = BarChart()
        bar_chart1.type = "col"
        bar_chart1.title = "Average Turns per Game by Model"
        bar_chart1.y_axis.title = "Average Turns"
        bar_chart1.x_axis.title = "Model"
        
        labels = Reference(ws3, min_col=1, min_row=2, max_row=len(all_models) + 1)
        data = Reference(ws3, min_col=5, min_row=1, max_row=len(all_models) + 1)
        bar_chart1.add_data(data, titles_from_data=True)
        bar_chart1.set_categories(labels)
        
        # Add data labels showing values on bars
        bar_chart1.dataLabels = DataLabelList()
        bar_chart1.dataLabels.showVal = True
        
        ws3.add_chart(bar_chart1, "H18")
        
        # === BAR CHART: Average Time ===
        bar_chart2 = BarChart()
        bar_chart2.type = "col"
        bar_chart2.title = "Average Time per Game by Model (seconds)"
        bar_chart2.y_axis.title = "Average Time (s)"
        bar_chart2.x_axis.title = "Model"
        
        labels = Reference(ws3, min_col=1, min_row=2, max_row=len(all_models) + 1)
        data = Reference(ws3, min_col=6, min_row=1, max_row=len(all_models) + 1)
        bar_chart2.add_data(data, titles_from_data=True)
        bar_chart2.set_categories(labels)
        
        # Add data labels showing values on bars
        bar_chart2.dataLabels = DataLabelList()
        bar_chart2.dataLabels.showVal = True
        
        ws3.add_chart(bar_chart2, "H34")
        
        # === BAR CHART: Win Rate ===
        bar_chart3 = BarChart()
        bar_chart3.type = "col"
        bar_chart3.title = "Win Rate by Model (%)"
        bar_chart3.y_axis.title = "Win Rate %"
        bar_chart3.x_axis.title = "Model"
        
        labels = Reference(ws3, min_col=1, min_row=2, max_row=len(all_models) + 1)
        data = Reference(ws3, min_col=4, min_row=1, max_row=len(all_models) + 1)
        bar_chart3.add_data(data, titles_from_data=True)
        bar_chart3.set_categories(labels)
        
        # Add data labels showing values on bars
        bar_chart3.dataLabels = DataLabelList()
        bar_chart3.dataLabels.showVal = True
        
        ws3.add_chart(bar_chart3, "P2")
        
        # Column widths
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 12
        ws3.column_dimensions['D'].width = 15
        ws3.column_dimensions['E'].width = 15
        ws3.column_dimensions['F'].width = 15
    
    def _calculate_avg_move_time(self, player: int) -> float:
        """Calculate average move time for a player"""
        player_times = [m['time_taken'] for m in self.move_times if m['player'] == player and m['is_valid']]
        return round(sum(player_times) / len(player_times), 4) if player_times else 0.0
    
    def close(self):
        try:
            self._f.close()
        except Exception:
            pass
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False